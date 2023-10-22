from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import MiUsuario
from .models import Fontanero
from .models import Predio
from .models import Anuncio
from .models import Sector
from .models import Multas
from .models import PagoAgua
from .models import Jornales
from django.contrib.auth import authenticate, login
from django.shortcuts import redirect
from django.contrib.auth import logout
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.hashers import make_password
from django.http import JsonResponse
from django.db.models import Q
import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from decimal import Decimal
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from paypalcheckoutsdk.core import PayPalHttpClient, SandboxEnvironment #eliminar al estar en produccion 

@login_required(login_url='task:login')
def es_administrador(user):
    return user.is_authenticated and user.is_staff

def es_administrador(user):
    return user.is_staff


def some_view(request):
    # ...
    next_url = '/restringida/'  # Reemplaza '/restringida/' con la URL adecuada de tu vista restringida
    return redirect('task:login')

@user_passes_test(es_administrador, login_url='task:vista_restringida')  # Redirige a 'vista_restringida' si no es administrador
def register(request):
    return render(request, 'registration/register.html')

def logout_view(request):
    logout(request)
    return redirect('task:login')

@login_required(login_url='task:login')
def vista_restringida(request):
    anuncios=Anuncio.objects.all()
    return render(request, 'Menu/menu.html',{'anuncios':anuncios})


@user_passes_test(es_administrador, login_url='task:vista_restringida')  # Redirige a 'vista_restringida' si no es administrador
def vista_restringida_Administrador(request):
    anuncios=Anuncio.objects.all()
    return render(request, 'Menu/menuAdministrador.html',{'anuncios':anuncios})

def login_view(request):
    if request.method == 'POST':
        dpi = request.POST['dpi']
        password = request.POST['password']
        user = authenticate(request, dpi=dpi, password=password)
        if user is not None:
            login(request, user)
            
            # Determinar a dónde redirigir al usuario según su rol
            if user.is_staff:  # Si es administrador
                return redirect('task:vista_restringida_Administrador')
            else:  # Si es usuario normal
                return redirect('task:vista_restringida')
    
    return render(request, 'registration/login.html')

@user_passes_test(es_administrador, login_url='task:vista_restringida')  # Redirige a 'vista_restringida' si no es administrador
def register_view(request):
    if request.method == 'POST':
        email = request.POST['email']
        nombres_apellidos = request.POST['nombres_apellidos']
        servicio = request.POST['servicio']
        telefono = request.POST['telefono']
        sector = request.POST.get('sector')
        sectors = Sector.objects.get(idSector=sector)
        titular = request.POST['titular']
        dpi = request.POST['dpi']
        password = request.POST['password']
        rol = request.POST['rol']  # Obtener el valor del campo de selección
        
        idsector= sectors.idSector
        # Crear el usuario
        user = MiUsuario.objects.create_user(email=email, dpi=dpi, password=password, nombres_apellidos=nombres_apellidos, servicio=servicio, telefono=telefono, sector_ubicacion_id=idsector, titular=titular)
        
        # Asignar permisos según el rol seleccionado
        if rol == 'administrador':
            user.is_staff = True
            user.is_superuser = True
        else:
            user.is_staff = False
            user.is_superuser = False
        
        user.save()  # Guardar los cambios
        
        
        # Redirigir a la vista restringida o a otra URL deseada
        return redirect('task:user_table')
    
    return render(request, 'registration/registration.html')

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def user_table_view(request):
    users = MiUsuario.objects.all()
    return render(request, 'Usuarios/usuarios.html', {'users': users})

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def Modify_user_view(request,idusuario):
    usuario= MiUsuario.objects.get(id=idusuario)
    print(usuario)
    return render(request,'Usuarios/modificar.html',{'usuario':usuario})

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def Modify_user(request,idusuario):
    usuario = MiUsuario.objects.get(id=idusuario)

    if request.method == 'POST':
        # Recuperar los datos del formulario
        email = request.POST['email']
        nombres_apellidos = request.POST['nombres_apellidos']
        servicio = request.POST['servicio']
        telefono = request.POST['telefono']
        direccion_sector = request.POST['direccion_sector']
        titular = request.POST['titular']
        dpi = request.POST['dpi']
        password = request.POST['password']

        # Crear un diccionario con los nuevos valores
        new_data = {
            'email': email,
            'nombres_apellidos': nombres_apellidos,
            'servicio': servicio,
            'telefono': telefono,
            'sector_ubicacion_id': direccion_sector,
            'titular': titular,
            'dpi': dpi,
        }

        if password:
            new_data['password'] = make_password(password)  # Hashear la contraseña

        # Actualizar los campos del usuario
        MiUsuario.objects.filter(id=idusuario).update(**new_data)

        # Redirigir a la página de detalles del usuario o a otra página deseada
        return redirect('task:user_table')

    return render(request, 'Usuarios/editar_usuario.html', {'usuario': usuario})

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def Delete_user(request,idusuario):
    usuario = MiUsuario.objects.get(id=idusuario)
    usuario.delete()
    return redirect('task:user_table')

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def View_fontanero(request):
    fontaneros = Fontanero.objects.all()
    return render(request, 'Fontaneros/fontaneros.html', {'fontaneros': fontaneros})

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def View_register_fontanero(request):
    return render(request, 'Fontaneros/register.html')

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def Register_fontanero(request):
    if request.method == 'POST':
        nombre = request.POST['nombre']
        estado = request.POST['estado']
        telefono = request.POST['telefono']
        tipo_pago = request.POST['tipo_pago']
        dpi = request.POST['dpi']
        
        fontanero = Fontanero.objects.create(nombre=nombre,telefono=telefono,dpi=dpi, estado=estado, tipo_pago=tipo_pago)
        
        fontanero.save()  
        fontaneros = Fontanero.objects.all()
        

        return redirect('task:view_fontanero') 
    
    return render(request, 'Fontaneros/register.html')

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def Modify_fontanero_view(request,idfontanero):
    fontanero= Fontanero.objects.get(idfontanero=idfontanero)
    return render(request,'Fontaneros/modificar.html',{'fontanero':fontanero})

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def Modify_fontanero(request, idfontanero):
    if request.method == 'POST':
        fontanero = Fontanero.objects.get(idfontanero=idfontanero)
        
        nombre = request.POST['nombre']
        estado = request.POST['estado']
        telefono = request.POST['telefono']
        tipo_pago = request.POST['tipo_pago']
        dpi = request.POST['dpi']
        

        fontanero.nombre = nombre
        fontanero.estado = estado
        fontanero.telefono = telefono
        fontanero.tipo_pago = tipo_pago
        fontanero.dpi = dpi
        
        fontanero.save()  
        
        return redirect('task:view_fontanero') 
    
@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def Delete_fontanero(request,idfontanero):
    fontanero = Fontanero.objects.get(idfontanero=idfontanero)
    fontanero.delete()
    return redirect('task:view_fontanero')

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def Delete_jornal(request,idjornal):
    jornal = Jornales.objects.get(id=idjornal)
    jornal.delete()
    return redirect('task:view_fontanero')


@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def editar_jornal_view(request,idjornal):
    jornal = Jornales.objects.get(id=idjornal)
    return render(request, 'Jornales/editarJornal.html', {'jornal':jornal})


@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def editar_jornal(request,idjornal):
    if request.method == 'POST':
        jornal = Jornales.objects.get(id=idjornal)
        idJornal= request.POST['idJornal']
        nombre = request.POST.get('nombre')
        telefono = request.POST['telefono']
        descripcion = request.POST['descripcion']
        sector = request.POST.get('sector')
        sectors = Sector.objects.get(idSector=sector)
        jornal.idJornal = idJornal
        jornal.nombre = nombre
        jornal.descripcion = descripcion
        jornal.telefono = telefono
        jornal.sector_id = sectors.idSector
        jornal.save()
        return redirect('task:jornales_view')
    

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def View_predios(request):
    predio = Predio.objects.all()
    return render(request, 'Predios/predio.html',{'predio':predio})

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def Delete_sector(request,idsector):
    sector = Sector.objects.get(idSector=idsector)
    sector.delete()
    return redirect('task:sector_view')

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def Delete_predio(request,idpredio):
    predio = Predio.objects.get(idpredio=idpredio)
    predio.delete()
    return redirect('task:view_predios')

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def Register_fontanero_view(request):
    return render(request, 'Predios/register.html')

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def search_users(request):
    search_term = request.GET.get('search')
    users = MiUsuario.objects.filter(Q(nombres_apellidos__icontains=search_term) | Q(email__icontains=search_term))
    user_list = [{'id': user.id, 'text': user.nombres_apellidos} for user in users]
    return JsonResponse(user_list, safe=False)

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def search_sector(request):
    try:
        search_term = request.GET.get('search')
        sectors = Sector.objects.filter(Q(nombre__icontains=search_term))
        sector_list = [{'id': sector.idSector, 'text': sector.nombre} for sector in sectors]
        return JsonResponse(sector_list, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)})
        
  
@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')    
def Register_predios(request):
    if request.method == 'POST':
        nombre = request.POST.get('usuario')
        ubicacion = request.POST['ubicacion']
        estado = request.POST['estado']
        usuario = None
        if nombre:  
            usuario = MiUsuario.objects.get(id=nombre)
            predio = Predio.objects.create(ubicacion=ubicacion,estado=estado,dueño_id=usuario.id)
            predio.save()
            return redirect('task:view_predios') 
        
        predio = Predio.objects.create(ubicacion=ubicacion,estado=estado,dueño_id=None)
        
        predio.save()  
        

        return redirect('task:view_predios')
    
    return render(request, 'Predios/register.html')


@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def Register_jornal(request):
    if request.method == 'POST':
        idJornal= request.POST['idJornal']
        nombre = request.POST.get('nombre')
        telefono = request.POST['telefono']
        descripcion = request.POST['descripcion']
        sector = request.POST.get('sector')
        sectors = Sector.objects.get(idSector=sector)
        jornal = Jornales.objects.create(idJornal=idJornal,nombre=nombre,descripcion=descripcion,telefono=telefono,sector_id=sectors.idSector)
        jornal.save()
        return redirect('task:jornales_view') 
        
    
    return render(request, 'Jornal/crearJornal.html')

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def Modify_predio_view(request,idpredio):
    predio= Predio.objects.get(idpredio=idpredio)
    return render(request, 'Predios/modificar.html', {'predio':predio})

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def Modify_predio(request, idpredio):
    if request.method == 'POST':
        predio = Predio.objects.get(idpredio=idpredio)
        nombre = request.POST.get('usuario')
        ubicacion = request.POST['ubicacion']
        estado = request.POST['estado']
        usuario = None
        if nombre:  
            predio.ubicacion = ubicacion
            predio.estado = estado
            predio.dueño_id = nombre
            predio.save()
            return redirect('task:view_predios') 
        

        predio.ubicacion = ubicacion
        predio.estado = estado
        predio.dueño_id = None
        
        predio.save()  
        
        return redirect('task:view_predios')
    
@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def Publicar_anuncio(request):
    user = request.user
    if request.method == 'POST':
        titulo = request.POST['titulo']
        contenido = request.POST['contenido']
        fecha_creacion = datetime.datetime.now()
        usuario_id = user.id
        
        anuncio = Anuncio.objects.create(titulo=titulo,contenido=contenido,fecha_creacion=fecha_creacion, usuario_id=usuario_id)
        
        anuncio.save()  
        

        return redirect('task:vista_restringida_Administrador')
    
    return render(request, 'Fontaneros/register.html')

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def Sector_view(request):
    sector = Sector.objects.all()
    return render(request, 'Sector/sector.html',{'sector':sector})

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def Create_sector(request):
    if request.method == 'POST':
        nombre = request.POST['nombre']
        descripcion = request.POST['descripcion']
        sector = Sector.objects.create(nombre=nombre,descripcion=descripcion)
        sector.save()
        return redirect('task:sector_view') 
    return render(request, 'Sector/sector.html')

from django.http import FileResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from io import BytesIO

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def some_view(request):
    # Crea un objeto BytesIO para guardar el PDF
    buffer = BytesIO()

    # Crea el PDF
    p = canvas.Canvas(buffer, pagesize=landscape(letter))

    # Aquí es donde se genera el contenido.
    data = MiUsuario.objects.all()  # Reemplaza 'MiUsuario' con el nombre de tu modelo

    # Dibuja cosas en el PDF. 
    x_offset = 50  # Posición inicial x
    y_offset = 750  # Posición inicial y
    for item in data:
        p.drawString(x_offset, y_offset, str(item.nombres_apellidos))  # Reemplaza 'nombres_apellidos' con el nombre de tu campo
        p.drawString(x_offset + 100, y_offset, str(item.servicio))  # Ajusta la posición x para cada campo
        p.drawString(x_offset + 200, y_offset, str(item.telefono))
        p.drawString(x_offset + 300, y_offset, str(item.sector_ubicacion_id))
        p.drawString(x_offset + 400, y_offset, str(item.titular))
        p.drawString(x_offset + 500, y_offset, str(item.dpi))
        y_offset -= 50  # Mueve la posición y para la siguiente línea

    # Cierra el objeto PDF limpiamente.
    p.showPage()
    p.save()

    # Crea un objeto de archivo PDF con el contenido del buffer
    buffer.seek(0)
    response = FileResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="somefilename.pdf"'

    return response

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def obtener_datos():
    datos = MiUsuario.objects.all()
    return datos

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def excel_report_usuario(request):
    # Obtener datos del modelo MiUsuario
    usuarios = MiUsuario.objects.all()

    # Crear un libro de trabajo de Excel y una hoja de trabajo
    workbook = Workbook()
    worksheet = workbook.active

    # Agregar encabezados
    worksheet['A1'] = 'Nombre y Apellido'
    worksheet['B1'] = 'Servicio'
    worksheet['C1'] = 'Teléfono'
    worksheet['D1'] = 'Dirección o Sector'
    worksheet['E1'] = 'Titular'
    worksheet['F1'] = 'DPI'

    # Agregar datos a la hoja de trabajo
    row_number = 2
    for usuario in usuarios:
        worksheet.cell(row=row_number, column=1, value=usuario.nombres_apellidos)
        worksheet.cell(row=row_number, column=2, value=usuario.servicio)
        worksheet.cell(row=row_number, column=3, value=usuario.telefono)
        worksheet.cell(row=row_number, column=4, value=usuario.sector_ubicacion_id)
        worksheet.cell(row=row_number, column=5, value=usuario.titular)
        worksheet.cell(row=row_number, column=6, value=usuario.dpi)
        row_number += 1

    # Crear una respuesta HTTP con el archivo Excel adjunto
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="reporte Usuarios.xlsx"'
    workbook.save(response)

    return response

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def excel_report_pagoagua(request):
    # Obtener datos de la tabla task_pagoagua
    pagos_agua = PagoAgua.objects.all()

    # Crear un libro de trabajo de Excel y una hoja de trabajo
    workbook = Workbook()
    worksheet = workbook.active

    # Agregar encabezados
    worksheet['A1'] = 'ID de Pago de Agua'
    worksheet['B1'] = 'Saldo'
    worksheet['C1'] = 'ID de Usuario'
    worksheet['D1'] = 'Mes'
    worksheet['E1'] = 'Pago Realizado'

    # Agregar datos a la hoja de trabajo
    row_number = 2
    for pago_agua in pagos_agua:
        worksheet.cell(row=row_number, column=1, value=pago_agua.idpagoagua)
        worksheet.cell(row=row_number, column=2, value=pago_agua.saldo)
        worksheet.cell(row=row_number, column=3, value=pago_agua.usuario.nombres_apellidos)
        worksheet.cell(row=row_number, column=4, value=pago_agua.mes)
        worksheet.cell(row=row_number, column=5, value=pago_agua.pago_realizado)
        row_number += 1

    # Crear una respuesta HTTP con el archivo Excel adjunto
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="reporte_pagoagua.xlsx"'
    workbook.save(response)

    return response

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def excel_report_fontanero(request):
    # Obtener datos del modelo TaskFontanero
    fontaneros = Fontanero.objects.all()

    # Crear un libro de trabajo de Excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Agregar encabezados
    headers = ['Nombre', 'Teléfono', 'DPI', 'Estado', 'Tipo de Pago']

    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        cell = worksheet['%s1' % column_letter]
        cell.value = header

    # Agregar datos a la hoja de trabajo
    for row_num, fontanero in enumerate(fontaneros, 2):
        row_data = [
            fontanero.nombre,
            fontanero.telefono,
            fontanero.dpi,
            fontanero.estado,
            fontanero.tipo_pago
        ]
        for col_num, cell_value in enumerate(row_data, 1):
            column_letter = get_column_letter(col_num)
            cell = worksheet.cell(row=row_num, column=col_num, value=cell_value)

    # Crear una respuesta HTTP con el archivo Excel adjunto
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="reporte Fontaneros.xlsx"'
    response.write(save_virtual_workbook(workbook))

    return response

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def excel_report_predio(request):
    # Obtener datos del modelo Predio
    predios = Predio.objects.all()

    # Crear un libro de trabajo de Excel y una hoja de trabajo
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Definir los encabezados según las columnas del modelo Predio
    headers = ['Ubicación', 'Estado', 'Dueño']

    # Agregar encabezados a la hoja de trabajo
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        cell = worksheet['%s1' % column_letter]
        cell.value = header

    # Agregar datos a la hoja de trabajo
    for row_num, predio in enumerate(predios, 2):
        row_data = [
            predio.ubicacion,
            predio.estado,
            str(predio.dueño.nombres_apellidos)  # Asegúrate de acceder al campo de nombre del dueño
        ]
        for col_num, cell_value in enumerate(row_data, 1):
            column_letter = get_column_letter(col_num)
            cell = worksheet.cell(row=row_num, column=col_num, value=cell_value)

    # Crear una respuesta HTTP con el archivo Excel adjunto
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="reporte Predios.xlsx"'
    response.write(save_virtual_workbook(workbook))

    return response

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def excel_report_jornales(request):
    # Obtener datos del modelo Jornales
    jornales = Jornales.objects.all()

    # Crear un libro de trabajo de Excel y una hoja de trabajo
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Definir los encabezados según las columnas del modelo Jornales
    headers = ['ID Jornal', 'Nombre', 'Teléfono', 'Descripción', 'Sector']

    # Agregar encabezados a la hoja de trabajo
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        cell = worksheet['%s1' % column_letter]
        cell.value = header

    # Agregar datos a la hoja de trabajo
    for row_num, jornal in enumerate(jornales, 2):
        row_data = [
            jornal.idJornal,
            jornal.nombre,
            jornal.telefono,
            jornal.descripcion,
            jornal.sector.nombre  # Accede al campo "nombre" del modelo Sector relacionado
        ]
        for col_num, cell_value in enumerate(row_data, 1):
            column_letter = get_column_letter(col_num)
            cell = worksheet.cell(row=row_num, column=col_num, value=cell_value)

    # Crear una respuesta HTTP con el archivo Excel adjunto
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="reporte Jornales.xlsx"'
    response.write(save_virtual_workbook(workbook))

    return response

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def excel_report_multas(request):
    # Obtener datos de la tabla Multas y relacionarlos con la tabla MiUsuario
    multas = Multas.objects.select_related('usuario')  # usuario es el campo que relaciona Multas con MiUsuario

    # Crear un libro de trabajo de Excel y una hoja de trabajo
    workbook = Workbook()
    worksheet = workbook.active

    # Agregar encabezados
    worksheet['A1'] = 'ID de Multa'
    worksheet['B1'] = 'Descripción'
    worksheet['C1'] = 'Saldo'
    worksheet['D1'] = 'Fecha de Creación'
    worksheet['E1'] = 'Fecha de Vencimiento'
    worksheet['F1'] = 'Nombre de Usuario'
    worksheet['G1'] = 'Pago Realizado'

    # Agregar datos a la hoja de trabajo
    row_number = 2
    for multa in multas:
        worksheet.cell(row=row_number, column=1, value=multa.idmultas)
        worksheet.cell(row=row_number, column=2, value=multa.descripcion)
        worksheet.cell(row=row_number, column=3, value=multa.saldo)
        worksheet.cell(row=row_number, column=4, value=multa.fecha_creacion.replace(tzinfo=None))  # Elimina la zona horaria
        worksheet.cell(row=row_number, column=5, value=multa.fecha_vencimiento.replace(tzinfo=None))
        worksheet.cell(row=row_number, column=6, value=multa.usuario.nombres_apellidos)  # Accede al nombre del usuario
        worksheet.cell(row=row_number, column=7, value=multa.pago_realizado)
        row_number += 1

    # Crear una respuesta HTTP con el archivo Excel adjunto
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="reporte_multas.xlsx"'
    workbook.save(response)

    return response
#como generar in filtro de busqueda por dpi, sector y nombre de usuario

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def search_users(request):
    search_term = request.GET.get('search')
    users = MiUsuario.objects.filter(Q(nombres_apellidos__icontains=search_term) | Q(email__icontains=search_term))
    user_list = [{'id': user.id, 'text': user.nombres_apellidos} for user in users]
    return JsonResponse(user_list, safe=False)

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def search_usuario(request):
     if request.method == 'POST':
        search = request.POST['search']
        users = MiUsuario.objects.filter(Q(nombres_apellidos__icontains=search) | Q(dpi__icontains=search))
        return render(request, 'Usuarios/usuarios.html',{'users':users})

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')     
def search_Fontaneros(request):
     if request.method == 'POST':
        search = request.POST['search']
        fontanero = Fontanero.objects.filter(Q(nombre__icontains=search) | Q(dpi__icontains=search) | Q(tipo_pago__icontains=search))
        return render(request, 'Fontaneros/fontaneros.html',{'fontaneros':fontanero})

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')     
def search_Jornal(request):
     if request.method == 'POST':
        search = request.POST['search']
        jornales = Jornales.objects.filter(Q(nombre__icontains=search) | Q(sector__nombre__icontains=search))
        return render(request, 'Jornales/jornales.html',{'jornales':jornales})


@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def search_Predios(request):
     if request.method == 'POST':
        search = request.POST['search']
        predios = Predio.objects.filter(Q(estado__icontains=search) | Q(dueño__nombres_apellidos__icontains=search))
        return render(request, 'Predios/predio.html',{'predio':predios})

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')     
def search_multas_pagos(request):
     if request.method == 'POST':
        search = request.POST['search']
        multas = Multas.objects.filter(Q(fecha_creacion__icontains=search) | Q(usuario__nombres_apellidos__icontains=search))
        multas= multas.filter(pago_realizado=True)
        return render(request, 'Pagos/pagos_multas.html',{'multa':multas})

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')     
def search_agua_pagos(request):
     if request.method == 'POST':
        search = request.POST['search']
        agua = PagoAgua.objects.filter(Q(mes__icontains=search) | Q(usuario__nombres_apellidos__icontains=search))
        agua= agua.filter(pago_realizado=True)
        return render(request, 'Pagos/pagos_servicioAgua.html',{'pagos_agua':agua})

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')     
def Create_view_multa(request, idusuario):
    usuario = MiUsuario.objects.get(id=idusuario)

    return render(request, 'Multas/create_multa.html', {'usuario': usuario})

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def create_multa(request,idusuario):
    if request.method == 'POST':
        descripcion = request.POST['descripcion']
        saldo_str = request.POST['saldo']
        saldo = Decimal(saldo_str)
        fecha_creacion = request.POST['fecha_creacion']
        fecha_vencimiento = request.POST['fecha_vencimiento']
        multa = Multas.objects.create(descripcion=descripcion,saldo=saldo,fecha_creacion=fecha_creacion,fecha_vencimiento=fecha_vencimiento,usuario_id=idusuario)
        multa.save()
        return redirect('task:view_user', idusuario=idusuario)
    
@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def View_User(request,idusuario):
    usuario = MiUsuario.objects.get(id=idusuario)
    multa = Multas.objects.filter(usuario_id=idusuario)
    agua = PagoAgua.objects.filter(usuario_id=idusuario)
    datos = {
        'multa': multa,
        'usuario': usuario,
        'agua': agua
    }
    return render(request, 'Usuarios/view_Multas.html',datos)

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def Delete_multa(request,idmulta):
    multa = Multas.objects.get(idmultas=idmulta)
    multa.delete()
    return redirect('task:view_user', idusuario=multa.usuario_id)

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def Delete_anuncio(request,idanuncio):
    anuncio = Anuncio.objects.get(id=idanuncio)
    anuncio.delete()
    return redirect('task:vista_restringida_Administrador')

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def pago_realizado_true(request, idmulta):
    multa = Multas.objects.get(idmultas=idmulta)
    #print(idmulta)
    if multa.pago_realizado:
        return redirect('task:view_user', idusuario=multa.usuario_id)

    
    saldo_multa = multa.saldo
    usuario = multa.usuario

    
    usuario.saldo -= saldo_multa

    
    multa.pago_realizado = True

    
    usuario.save()
    multa.save()

    return redirect('task:view_user', idusuario=usuario.id)

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def pagos_views(request):
    multas_pagadas = Multas.objects.filter(pago_realizado=True)
    datos = {
        'multa': multas_pagadas
    }

    return render(request, 'Pagos/pagos_multas.html',datos)

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def pagos_views_agua(request):
    pagos_agua_realizados = PagoAgua.objects.filter(pago_realizado=True)
    datos = {
        'pagos_agua': pagos_agua_realizados,
    }

    return render(request, 'Pagos/pagos_servicioAgua.html',datos)

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def pago_realizado_agua_true(request, idpagoagua):
    agua = PagoAgua.objects.get(idpagoagua=idpagoagua)

    if agua.pago_realizado:
        # Si el pago ya se ha realizado, no hagas nada y redirige
        return redirect('task:view_user', idusuario=agua.usuario_id)

    # Obtén el saldo de la multa y el usuario asociado
    saldo_multa = agua.saldo
    usuario = agua.usuario

    # Resta el saldo de la multa al saldo del usuario
    usuario.saldo -= saldo_multa

    # Marca la multa como pagada
    agua.pago_realizado = True

    # Guarda los cambios en la base de datos
    usuario.save()
    agua.save()

    return redirect('task:view_user', idusuario=usuario.id)

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def jornales_view(request):
    jornales = Jornales.objects.all()
    return render(request, 'Jornales/jornales.html',{'jornales':jornales})

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def crear_jornales_view(request):
    return render(request, 'Jornales/crearJornal.html')


#vista usuario normal
@login_required(login_url='task:login')
def pendientes_usuario(request):
    usuario_id=request.user.id
    multa = Multas.objects.filter(usuario_id=usuario_id)
    agua = PagoAgua.objects.filter(usuario_id=usuario_id)
    datos = {
        'multa': multa,
        'usuario': usuario_id,
        'agua': agua,
        
    }

    return render(request, 'Usuarios_normales/pendientes.html',datos)

@login_required(login_url='task:login')
def contactanos_view(request):
    return render(request, 'Usuarios_normales/Contactar.html')

@login_required(login_url='task:login')
@user_passes_test(es_administrador, login_url='task:vista_restringida')
def search_fontanero(request):
     if request.method == 'POST':
        search = request.POST['search']
        fontanero = Fontanero.objects.filter(Q(nombres_apellidos__icontains=search) | Q(dpi__icontains=search))
        return render(request, 'Usuarios/usuarios.html',{'users':users})

@login_required(login_url='task:login')   
def pendientes_multas(request):
    user = request.user
    multa = Multas.objects.filter(usuario_id=user.id)
    datos = {
        'multa': multa,
        'usuario': user.id,
    }
    return render(request, 'Usuarios_normales/pendientes_multas.html',datos) 
     
     
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
import json
from paypalcheckoutsdk.core import PayPalHttpClient, SandboxEnvironment
from paypalcheckoutsdk.orders import OrdersCreateRequest, OrdersCaptureRequest

@login_required(login_url='task:login')
def pago_completado_view(request):
    data = json.loads(request.body.decode("utf-8"))
    print(data)
    idagua = data.get('idagua')
    print(f'id de pago de agua: {idagua}')
    pagoagua=PagoAgua.objects.get(idpagoagua=idagua)
    saldo_multa = pagoagua.saldo
    usuario = pagoagua.usuario
    usuario.saldo -= saldo_multa

    pagoagua.pago_realizado=True
    usuario.save()
    pagoagua.save()
    
    
    return render(request, 'Usuarios_normales/pendientes.html')

@login_required(login_url='task:login')
def pagomulta_completado_view(request):
    data = json.loads(request.body.decode("utf-8"))
    print(data)
    idmultas = data.get('idmultas')
    print(f'id de pago de multa: {idmultas}')
    pagomulta=Multas.objects.get(idmultas=idmultas)
    saldo_multa = pagomulta.saldo
    usuario = pagomulta.usuario
    usuario.saldo -= saldo_multa
    pagomulta.pago_realizado=True
    usuario.save()
    pagomulta.save()
    
    
    return render(request, 'Usuarios_normales/pendientes_multas.html')


    



